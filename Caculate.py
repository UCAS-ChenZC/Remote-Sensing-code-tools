from openpyxl import load_workbook

# 加载工作簿
wb = load_workbook('/home/solid/results.ods')  # 替换为你的文件名
ws = wb.active  # 或者使用 wb.get_sheet_by_name('Sheet1') 来获取特定的工作表

# 假设数据从第2行开始，第1行是标题行
for row in ws.iter_rows(min_row=2, values_only=True):
    precision = row[4]  # E列的值
    recall = row[5]     # F列的值

    # 检查值是否为数字，避免除以零的错误
    if isinstance(precision, (int, float)) and isinstance(recall, (int, float)):
        f1_score = (2 * precision * recall) / (precision + recall)
        # 将F1分数写入G列
        ws.cell(row=row[0], column=7, value=f1_score)

# 保存工作簿
wb.save('/home/solid/test.ods')  # 保存为新的文件，以避免覆盖原始文件
